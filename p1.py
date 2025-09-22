import pandas as pd
import numpy as np
import re
from tkinter import Tk, filedialog
from openpyxl import load_workbook
import os
import openpyxl

from tkinter import Tk, filedialog

import tkinter as tk
from tkinter import filedialog, simpledialog, messagebox

from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime


Tk().withdraw()

print("CP-ESCOA Cost extract Weekly reconcilition Final Jul 2025:")
main_path = filedialog.askopenfilename(title="CP-ESCOA Cost extract Weekly reconcilition Final Jul 2025")

# Validate the selected file
if not main_path:
    print("No file selected. Exiting...")
    exit()

if not os.path.exists(main_path):
    print(f"Selected file does not exist: {main_path}")
    exit()

print(f"Selected file: {main_path}")
print(f"File size: {os.path.getsize(main_path)} bytes")

# Create a backup of the original file before making any changes
def create_backup(file_path):
    """Create a backup of the original file with timestamp"""
    try:
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = f"{os.path.splitext(file_path)[0]}_backup_{timestamp}.xlsx"
        
        # Copy the file
        import shutil
        shutil.copy2(file_path, backup_path)
        print(f"Backup created: {backup_path}")
        return backup_path
    except Exception as e:
        print(f"Warning: Could not create backup: {str(e)}")
        return None

# Create backup before processing
backup_file = create_backup(main_path)



# Validate that required sheets exist in the file
try:
    # Check if required sheets exist
    workbook = load_workbook(main_path)
    available_sheets = workbook.sheetnames
    print(f"Available sheets in file: {available_sheets}")
    
    required_sheets = ["Cost extract report", "GL"]
    missing_sheets = [sheet for sheet in required_sheets if sheet not in available_sheets]
    
    if missing_sheets:
        print(f"Error: Required sheets not found: {missing_sheets}")
        print("Please ensure the file contains both 'Cost extract report' and 'GL' sheets.")
        exit()
    
    print("All required sheets found. Proceeding with data processing...")
    
except Exception as e:
    print(f"Error reading file structure: {str(e)}")
    exit()

# Read the data from the validated sheets
Cost_extract = pd.read_excel(main_path, sheet_name="Cost extract report")
GL = pd.read_excel(main_path, sheet_name="GL")

print(f"Successfully loaded data:")
print(f"  - Cost extract report: {len(Cost_extract)} rows")
print(f"  - GL: {len(GL)} rows")

# Create two separate tables with unique Concatenate values and summed amounts

# Table 1: GL data with unique Concatenate and summed Entered Amount
print("\nProcessing GL data...")
GL_unique = GL.groupby('Concatenate')['Entered Amount'].sum().reset_index()
GL_unique.columns = ['Concatenate', 'Total Entered Amount']

# Table 2: Cost extract data with unique Concatenate and summed Invoice Amount
print("Processing Cost extract data...")
Cost_extract_unique = Cost_extract.groupby('Concatenate')['Invoice Amount'].sum().reset_index()
Cost_extract_unique.columns = ['Concatenate', 'Total Invoice Amount']

# Display the results
print("\n=== GL Table (Unique Concatenate with Summed Entered Amount) ===")
print(GL_unique.head(10))  # Show first 10 rows
print(f"\nTotal unique Concatenate values in GL: {len(GL_unique)}")

print("\n=== Cost Extract Table (Unique Concatenate with Summed Invoice Amount) ===")
print(Cost_extract_unique.head(10))  # Show first 10 rows
print(f"\nTotal unique Concatenate values in Cost Extract: {len(Cost_extract_unique)}")

# Create merged dataset with Concatenate as key and both amounts side by side
print("\nCreating merged dataset...")
merged_dataset = pd.merge(GL_unique, Cost_extract_unique, on='Concatenate', how='outer')

# Fill NaN values with 0 for better readability
merged_dataset = merged_dataset.fillna(0)

# Display the merged dataset
print("\n=== Merged Dataset (Concatenate as Key with Both Amounts) ===")
print(merged_dataset.head(10))  # Show first 10 rows
print(f"\nTotal rows in merged dataset: {len(merged_dataset)}")
print(f"Columns in merged dataset: {list(merged_dataset.columns)}")

# Rename columns in the merged dataset
print("\nRenaming columns in merged dataset...")
merged_dataset = merged_dataset.rename(columns={
    'Concatenate': 'Row Label',
    'Total Entered Amount': 'GL',
    'Total Invoice Amount': 'Cost_extract'
})

# Display the renamed merged dataset
print("\n=== Renamed Merged Dataset ===")
print(merged_dataset.head(10))
print(f"Columns in renamed merged dataset: {list(merged_dataset.columns)}")

# Create a new Variance column with sum of GL and Cost_extract
print("\nCreating Variance column...")
merged_dataset['Variance'] = merged_dataset['GL'] + merged_dataset['Cost_extract']

# Check if absolute variance is less than 1, then mark as zero
print("Checking variance values and marking small variances as zero...")
merged_dataset['Variance'] = merged_dataset['Variance'].apply(lambda x: 0 if abs(x) < 1 else x)

# Display the final dataset with Variance column
print("\n=== Final Dataset with Variance Column ===")
print(merged_dataset.head(10))
print(f"Columns in final dataset: {list(merged_dataset.columns)}")

# Save the final dataset to the existing Excel file without disturbing other sheets
print("\nSaving final dataset to existing Excel file...")
output_filename = main_path

# Load the existing workbook to check existing sheets
workbook = load_workbook(output_filename)
existing_sheets = set(workbook.sheetnames)
print(f"Existing sheets in file: {list(existing_sheets)}")

# Check if Reconciliation_Summary sheet already exists
if 'Reconciliation_Summary' in existing_sheets:
    print("Warning: Reconciliation_Summary sheet already exists!")
    print("To preserve existing data, we will create a new sheet with timestamp.")
    from datetime import datetime
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    sheet_name = f'Reconciliation_Summary_{timestamp}'
else:
    sheet_name = 'Reconciliation_Summary'

print(f"Creating sheet: {sheet_name}")

# Use ExcelWriter with mode='a' to append to existing file
# Use 'new' to ensure we don't overwrite anything
with pd.ExcelWriter(output_filename, mode='a', engine='openpyxl', if_sheet_exists='new') as writer:
    merged_dataset.to_excel(writer, index=False, sheet_name=sheet_name)

print(f"Final dataset saved successfully to: {output_filename}")
print(f"File contains {len(merged_dataset)} rows and {len(merged_dataset.columns)} columns")

# Show final status of the file
final_workbook = load_workbook(output_filename)
final_sheets = set(final_workbook.sheetnames)
print(f"\nFinal file status:")
print(f"Total sheets in file: {len(final_sheets)}")
print(f"All sheets: {list(final_sheets)}")
print(f"New sheet created: {sheet_name}")
print("✓ All existing data has been preserved")
if backup_file:
    print(f"✓ Original file backed up as: {backup_file}")

# Now you have three datasets:
# 1. GL_unique - GL data with unique Concatenate and summed Entered Amount
# 2. Cost_extract_unique - Cost extract data with unique Concatenate and summed Invoice Amount  
# 3. merged_dataset - Combined dataset with Row Label as key, GL, Cost_extract amounts side by side, and Variance column
# 4. New Excel file: Cost_Extract_Reconciliation_Summary.xlsx with the final dataset


def normalize_date(date_value):
    """
    Convert various date formats to datetime object for consistent comparison.
    Handles pandas Timestamp, string dates, and datetime objects.
    """
    if pd.isna(date_value) or date_value is None:
        return None
    
    # If it's already a datetime object, return as is
    if isinstance(date_value, datetime):
        return date_value
    
    # If it's a pandas Timestamp, convert to datetime
    if hasattr(date_value, 'to_pydatetime'):
        return date_value.to_pydatetime()
    
    # If it's a string, try to parse it
    if isinstance(date_value, str):
        try:
            # Try common date formats
            for fmt in ['%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S']:
                try:
                    return datetime.strptime(date_value, fmt)
                except ValueError:
                    continue
            # If all formats fail, try pandas to_datetime
            return pd.to_datetime(date_value).to_pydatetime()
        except:
            return None
    
    return None

def create_variance_sheets(main_file_path=None):
    """
    Read the main Excel file and create new sheets for each Row Label
    where Variance is not 0. Each sheet will be named after the corresponding Row Label.
    Uses the same main file for both source data and output.
    """
    
    # Use the main file path if provided, otherwise try to find it
    if main_file_path:
        input_file = main_file_path
        source_file = main_file_path
    else:
        # Try to find the main file in current directory
        input_file = "Cost_Extract_Reconciliation_Summary.xlsx"
        source_file = "CP-ESCOA Cost extract Weekly reconcilition Final Jul 2025.xlsx"
    
    # Check if input file exists
    if not os.path.exists(input_file):
        print(f"Error: {input_file} not found!")
        print("Please ensure the file exists in the current directory.")
        return
    
    # Check if source file exists
    if not os.path.exists(source_file):
        print(f"Error: {source_file} not found!")
        print("Please ensure the source file exists in the current directory.")
        return
    
    try:
        # Load the workbook to check for reconciliation sheets first
        workbook = load_workbook(input_file)
        existing_sheets = set(workbook.sheetnames)
        print(f"Existing sheets in workbook: {list(existing_sheets)}")
        
        # Check if Reconciliation_Summary sheet exists (or any variation with timestamp)
        reconciliation_sheets = [sheet for sheet in existing_sheets if sheet.startswith('Reconciliation_Summary')]
        if not reconciliation_sheets:
            print("Warning: No Reconciliation_Summary sheet found in the file!")
            print("Please run the main reconciliation process first.")
            return
        
        # Use the most recent Reconciliation_Summary sheet if multiple exist
        if len(reconciliation_sheets) > 1:
            print(f"Found multiple reconciliation sheets: {reconciliation_sheets}")
            # Sort by timestamp if available, otherwise use the first one
            reconciliation_sheets.sort(reverse=True)
        
        reconciliation_sheet_name = reconciliation_sheets[0]
        print(f"Using reconciliation sheet: {reconciliation_sheet_name}")
        
        # Read the reconciliation summary
        print(f"Reading {input_file}...")
        reconciliation_data = pd.read_excel(input_file, sheet_name=reconciliation_sheet_name)
        
        print(f"Successfully read {len(reconciliation_data)} rows from {reconciliation_sheet_name} sheet")
        
        # Filter rows where Variance is not 0
        non_zero_variance = reconciliation_data[reconciliation_data['Variance'] != 0]
        
        if len(non_zero_variance) == 0:
            print("No rows found with non-zero Variance. All reconciliations are balanced.")
            return
        
        print(f"Found {len(non_zero_variance)} rows with non-zero Variance")
        
        # Read source data for tables
        print("Reading source data for tables...")
        try:
            # Read GL sheet data
            gl_data = pd.read_excel(source_file, sheet_name="GL")
            print(f"Successfully read GL data: {len(gl_data)} rows")
            
            # Read Cost extract report data
            cost_extract_data = pd.read_excel(source_file, sheet_name="Cost extract report")
            print(f"Successfully read Cost extract report data: {len(cost_extract_data)} rows")
            
        except Exception as e:
            print(f"Error reading source data: {str(e)}")
            return
        
        # Process each row with non-zero variance
        sheets_created = 0
        for index, row in non_zero_variance.iterrows():
            row_label = str(row['Row Label'])
            
            # Clean the row label to make it a valid sheet name
            # Excel sheet names cannot contain: \ / ? * [ ]
            sheet_name = clean_sheet_name(row_label)
            
            # Check if sheet already exists
            if sheet_name in existing_sheets:
                print(f"Warning: Sheet '{sheet_name}' already exists.")
                # Create a unique sheet name with timestamp
                from datetime import datetime
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                original_sheet_name = sheet_name
                sheet_name = f"{original_sheet_name}_{timestamp}"
                print(f"Creating new sheet with unique name: {sheet_name}")
            
            try:
                # Create new sheet
                new_sheet = workbook.create_sheet(title=sheet_name)
                
                # Create tables in the sheet
                create_tables_in_sheet(new_sheet, gl_data, cost_extract_data, row_label)
                
                sheets_created += 1
                print(f"Created sheet with tables: {sheet_name}")
                
            except Exception as e:
                print(f"Error creating sheet '{sheet_name}': {str(e)}")
                continue
        
        # Save the workbook with new sheets
        print(f"\nSaving workbook with {sheets_created} new sheets...")
        workbook.save(input_file)
        
        print(f"Successfully created {sheets_created} new sheets with tables in {input_file}")
        print("Each sheet contains GL and IBS tables with aggregated data.")
        
        # Display summary
        print(f"\nSummary:")
        print(f"- Total rows processed: {len(reconciliation_data)}")
        print(f"- Rows with non-zero variance: {len(non_zero_variance)}")
        print(f"- New sheets created: {sheets_created}")
        
    except Exception as e:
        print(f"Error processing file: {str(e)}")

def create_tables_in_sheet(sheet, gl_data, cost_extract_data, row_label):
    """
    Create GL and IBS tables in the given sheet with aggregated data.
    """
    try:
        # Filter data based on row_label (Concatenate field)
        # For GL data - filter by Concatenate field
        gl_filtered = gl_data[gl_data['Concatenate'] == row_label]
        
        # For Cost extract data - filter by Concatenate field  
        cost_extract_filtered = cost_extract_data[cost_extract_data['Concatenate'] == row_label]
        
        # Create GL Table (left side)
        # Aggregate GL data by Posted Date
        if len(gl_filtered) > 0:
            gl_aggregated = gl_filtered.groupby('Posted Date')['Entered Amount'].sum().reset_index()
            # Convert dates to datetime for proper sorting
            gl_aggregated['Posted Date'] = pd.to_datetime(gl_aggregated['Posted Date'], errors='coerce')
            gl_aggregated = gl_aggregated.sort_values('Posted Date')
        else:
            gl_aggregated = pd.DataFrame(columns=['Posted Date', 'Entered Amount'])
        
        # Create IBS Table (right side)
        # Aggregate Cost extract data by Invoice Date
        if len(cost_extract_filtered) > 0:
            ibs_aggregated = cost_extract_filtered.groupby('Invoice Date')['Invoice Amount'].sum().reset_index()
            # Convert dates to datetime for proper sorting
            ibs_aggregated['Invoice Date'] = pd.to_datetime(ibs_aggregated['Invoice Date'], errors='coerce')
            ibs_aggregated = ibs_aggregated.sort_values('Invoice Date')
        else:
            ibs_aggregated = pd.DataFrame(columns=['Invoice Date', 'Invoice Amount'])
        
        # Position tables side by side
        # GL Table starts at A1
        sheet['A1'] = 'GL'
        sheet['A1'].font = Font(bold=True, size=14)
        sheet['A2'] = 'Posted Date'
        sheet['B2'] = 'Entered Amount'
        
        # Style headers
        for cell in ['A2', 'B2']:
            sheet[cell].font = Font(bold=True)
            sheet[cell].alignment = Alignment(horizontal='center')
        
        # Fill GL table data
        for idx, row in gl_aggregated.iterrows():
            row_num = idx + 3
            sheet[f'A{row_num}'] = row['Posted Date']
            sheet[f'B{row_num}'] = row['Entered Amount']
        
        # IBS Table starts at D1 (right side)
        sheet['D1'] = 'IBS'
        sheet['D1'].font = Font(bold=True, size=14)
        sheet['D2'] = 'Invoice Date'
        sheet['E2'] = 'Invoice Amount'
        
        # Style headers
        for cell in ['D2', 'E2']:
            sheet[cell].font = Font(bold=True)
            sheet[cell].alignment = Alignment(horizontal='center')
        
        # Fill IBS table data
        for idx, row in ibs_aggregated.iterrows():
            row_num = idx + 3
            sheet[f'D{row_num}'] = row['Invoice Date']
            sheet[f'E{row_num}'] = row['Invoice Amount']
        
        # Add borders to tables
        add_table_borders(sheet, len(gl_aggregated), len(ibs_aggregated))
        
        # Auto-adjust column widths
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            sheet.column_dimensions[col].width = 15
        
        # Compare and mark unmatched values
        unmatched_data = compare_and_mark_unmatched(sheet, gl_aggregated, ibs_aggregated)
        
        # Create detailed unmatched tables below the main tables
        create_detailed_unmatched_tables(sheet, gl_data, cost_extract_data, row_label, gl_aggregated, ibs_aggregated)
        
    except Exception as e:
        print(f"Error creating tables in sheet: {str(e)}")

def compare_and_mark_unmatched(sheet, gl_aggregated, ibs_aggregated):
    """
    Compare amounts between GL and IBS tables and mark unmatched values.
    Treats all amounts as absolute values for comparison.
    Also checks if IBS dates go beyond the latest GL date and marks them as 'TIME'.
    """
    try:
        # Add column headers for the new status columns
        sheet['C2'] = 'Status'
        sheet['F2'] = 'Status'
        sheet['C2'].font = Font(bold=True)
        sheet['F2'].font = Font(bold=True)
        sheet['C2'].alignment = Alignment(horizontal='center')
        sheet['F2'].alignment = Alignment(horizontal='center')
        
        # Get all amounts from both tables as absolute values
        gl_amounts = []
        ibs_amounts = []
        
        # Collect GL amounts and dates
        gl_dates = []
        for idx, row in gl_aggregated.iterrows():
            amount = abs(float(row['Entered Amount']))
            gl_amounts.append((idx, amount))
            normalized_date = normalize_date(row['Posted Date'])
            gl_dates.append(normalized_date)
        
        # Collect IBS amounts and dates
        ibs_dates = []
        for idx, row in ibs_aggregated.iterrows():
            amount = abs(float(row['Invoice Amount']))
            ibs_amounts.append((idx, amount))
            normalized_date = normalize_date(row['Invoice Date'])
            ibs_dates.append(normalized_date)
        
        # Find the latest GL date for comparison (filter out None values)
        latest_gl_date = None
        valid_gl_dates = [d for d in gl_dates if d is not None]
        if valid_gl_dates:
            latest_gl_date = max(valid_gl_dates)
        
        # Create sets of unique amounts for comparison
        gl_unique_amounts = set(amount for _, amount in gl_amounts)
        ibs_unique_amounts = set(amount for _, amount in ibs_amounts)
        
        # Find unmatched amounts
        gl_unmatched = gl_unique_amounts - ibs_unique_amounts
        ibs_unmatched = ibs_unique_amounts - gl_unique_amounts
        
        # Mark unmatched GL amounts in column C
        for idx, amount in gl_amounts:
            if amount in gl_unmatched:
                sheet[f'C{idx + 3}'] = 'UNMATCHED'
                sheet[f'C{idx + 3}'].font = Font(color='FF0000', bold=True)  # Red color
            else:
                sheet[f'C{idx + 3}'] = 'MATCHED'
                sheet[f'C{idx + 3}'].font = Font(color='008000', bold=True)  # Green color
        
        # Mark IBS amounts in column F with date comparison
        for idx, (amount, date) in enumerate(zip(ibs_amounts, ibs_dates)):
            # Check if IBS date goes beyond the latest GL date
            if latest_gl_date and date is not None and date > latest_gl_date:
                sheet[f'F{idx + 3}'] = 'TIME'
                sheet[f'F{idx + 3}'].font = Font(color='FF8C00', bold=True)  # Orange color for TIME
            elif amount[1] in ibs_unmatched:
                sheet[f'F{idx + 3}'] = 'UNMATCHED'
                sheet[f'F{idx + 3}'].font = Font(color='FF0000', bold=True)  # Red color
            else:
                sheet[f'F{idx + 3}'] = 'MATCHED'
                sheet[f'F{idx + 3}'].font = Font(color='008000', bold=True)  # Green color
        
        # Count TIME entries (IBS dates beyond latest GL date)
        time_entries = 0
        for date in ibs_dates:
            if latest_gl_date and date is not None and date > latest_gl_date:
                time_entries += 1
        
        # Add summary below the tables
        summary_row = max(len(gl_aggregated), len(ibs_aggregated)) + 5
        
        if gl_unmatched or ibs_unmatched or time_entries > 0:
            sheet[f'A{summary_row}'] = 'RECONCILIATION SUMMARY:'
            sheet[f'A{summary_row}'].font = Font(bold=True, size=12)
            summary_row += 1
            
            sheet[f'A{summary_row}'] = f'GL Total Rows: {len(gl_aggregated)}'
            summary_row += 1
            sheet[f'A{summary_row}'] = f'IBS Total Rows: {len(ibs_aggregated)}'
            summary_row += 1
            summary_row += 1
            
            if latest_gl_date:
                sheet[f'A{summary_row}'] = f'Latest GL Date: {latest_gl_date.strftime("%Y-%m-%d")}'
                summary_row += 1
                summary_row += 1
            
            if gl_unmatched:
                sheet[f'A{summary_row}'] = f'GL Unmatched Amounts: {len(gl_unmatched)}'
                sheet[f'A{summary_row}'].font = Font(color='FF0000', bold=True)
                summary_row += 1
                
            if ibs_unmatched:
                sheet[f'A{summary_row}'] = f'IBS Unmatched Amounts: {len(ibs_unmatched)}'
                sheet[f'A{summary_row}'].font = Font(color='FF0000', bold=True)
                summary_row += 1
            
            if time_entries > 0:
                sheet[f'A{summary_row}'] = f'IBS Time Issues (dates beyond GL): {time_entries}'
                sheet[f'A{summary_row}'].font = Font(color='FF8C00', bold=True)
                summary_row += 1
        else:
            sheet[f'A{summary_row}'] = 'ALL AMOUNTS MATCHED ✓'
            sheet[f'A{summary_row}'].font = Font(bold=True, color='008000', size=12)
        
        # Return unmatched data for detailed table creation
        return {
            'gl_unmatched_amounts': gl_unmatched,
            'ibs_unmatched_amounts': ibs_unmatched,
            'time_entries': time_entries,
            'latest_gl_date': latest_gl_date
        }
        
    except Exception as e:
        print(f"Error in comparison: {str(e)}")
        return {
            'gl_unmatched_amounts': set(),
            'ibs_unmatched_amounts': set(),
            'time_entries': 0,
            'latest_gl_date': None
        }

def get_unmatched_dates_from_status(sheet, aggregated_data, status_column):
    """
    Get dates that have UNMATCHED status from the Status column.
    """
    unmatched_dates = []
    try:
        for idx, row in aggregated_data.iterrows():
            # Check the status in the specified column (C for GL, F for IBS)
            status_cell = sheet[f'{status_column}{idx + 3}']  # +3 because data starts at row 3
            if status_cell.value == 'UNMATCHED':
                # Get the date from the first column (A for GL, D for IBS)
                date_cell = sheet[f'A{idx + 3}'] if status_column == 'C' else sheet[f'D{idx + 3}']
                if date_cell.value:
                    unmatched_dates.append(date_cell.value)
    except Exception as e:
        print(f"Error getting unmatched dates from status: {str(e)}")
    
    return unmatched_dates

def create_detailed_unmatched_tables(sheet, gl_data, cost_extract_data, row_label, gl_aggregated, ibs_aggregated):
    """
    Create detailed tables below the main GL and IBS tables for unmatched data.
    Checks Status column for UNMATCHED entries and creates tables 5-6 lines below main tables.
    Also performs validation and adds validation summary.
    """
    try:
        # Calculate starting row for detailed tables (5-6 lines below main tables)
        main_table_end_row = max(len(gl_aggregated), len(ibs_aggregated)) + 3  # +3 for headers
        detailed_table_start_row = main_table_end_row + 6  # 6 lines below main tables
        
        # Filter data for the specific row_label
        gl_filtered = gl_data[gl_data['Concatenate'] == row_label]
        cost_extract_filtered = cost_extract_data[cost_extract_data['Concatenate'] == row_label]
        
        # Get unmatched dates from the main tables by checking Status column
        gl_unmatched_dates = get_unmatched_dates_from_status(sheet, gl_aggregated, 'C')  # Column C for GL status
        ibs_unmatched_dates = get_unmatched_dates_from_status(sheet, ibs_aggregated, 'F')  # Column F for IBS status
        
        # Create detailed GL table for unmatched data
        gl_detailed_data, gl_data_start_row = create_detailed_gl_table(sheet, gl_filtered, gl_unmatched_dates, detailed_table_start_row)
        
        # Create detailed IBS table for unmatched data (side by side)
        ibs_detailed_data, ibs_data_start_row = create_detailed_ibs_table(sheet, cost_extract_filtered, ibs_unmatched_dates, detailed_table_start_row)
        
        # Cross-compare detailed tables and update status
        cross_compare_detailed_tables(sheet, gl_detailed_data, ibs_detailed_data, gl_data_start_row, ibs_data_start_row)
        
        # Perform validation on the detailed unmatched data
        variable_A, variable_B, validation_summary, cost_extract_unmatched = validate_unmatched_data(
            sheet, gl_detailed_data, ibs_detailed_data, cost_extract_data, row_label
        )
        
        # Compare variables A and B and create Table 1 if they are equal
        table1_created = compare_variables_and_create_table1(sheet, variable_A, variable_B, cost_extract_unmatched)
        
        # Create RTS table for invoices with unmatched status not present in both tables
        unmatched_not_in_both = find_unmatched_invoices_not_in_both_tables(gl_detailed_data, ibs_detailed_data, sheet)
        rts_data = get_entered_net_from_gl(gl_data, unmatched_not_in_both, row_label)
        create_rts_table(sheet, rts_data, start_row=1)
        
        # Calculate starting row for validation summary (below detailed tables)
        detailed_table_end_row = max(
            gl_data_start_row + len(gl_detailed_data) if len(gl_detailed_data) > 0 else detailed_table_start_row,
            ibs_data_start_row + len(ibs_detailed_data) if len(ibs_detailed_data) > 0 else detailed_table_start_row
        )
        validation_start_row = detailed_table_end_row + 3  # 3 lines below detailed tables
        
        # Add validation summary to the sheet
        add_validation_summary_to_sheet(sheet, variable_A, variable_B, validation_summary, validation_start_row)
        
        # Store validation results for potential future use
        sheet._validation_results = {
            'variable_A': variable_A,
            'variable_B': variable_B,
            'validation_summary': validation_summary,
            'table1_created': table1_created,
            'cost_extract_unmatched': cost_extract_unmatched,
            'rts_data': rts_data,
            'unmatched_not_in_both': unmatched_not_in_both
        }
        
    except Exception as e:
        print(f"Error creating detailed unmatched tables: {str(e)}")

def create_detailed_gl_table(sheet, gl_filtered, gl_unmatched_dates, start_row):
    """
    Create detailed GL table below main GL table.
    Shows individual line items only for unmatched dates.
    """
    try:
        # Filter GL data to show only line items for unmatched dates
        if gl_unmatched_dates:
            # Convert dates to datetime for comparison
            gl_filtered_copy = gl_filtered.copy()
            gl_filtered_copy['Posted Date'] = pd.to_datetime(gl_filtered_copy['Posted Date'], errors='coerce')
            
            # Filter to show only line items for unmatched dates
            gl_detailed = gl_filtered_copy[gl_filtered_copy['Posted Date'].isin(gl_unmatched_dates)]
        else:
            gl_detailed = pd.DataFrame()  # Empty dataframe if no unmatched dates
        
        if len(gl_detailed) == 0:
            # Still create the table header even if no data
            sheet[f'A{start_row}'] = 'DETAILED GL UNMATCHED DATA'
            sheet[f'A{start_row}'].font = Font(bold=True, size=12, color='FF0000')
            start_row += 1
            sheet[f'A{start_row}'] = 'Invoice Number'
            sheet[f'B{start_row}'] = 'Entered Amount'
            start_row += 1
            sheet[f'A{start_row}'] = 'No unmatched data found'
            return
        
        # Sort by Posted Date first, then by Invoice Number
        gl_detailed = gl_detailed.copy()
        gl_detailed['Posted Date'] = pd.to_datetime(gl_detailed['Posted Date'], errors='coerce')
        gl_detailed = gl_detailed.sort_values(['Posted Date', 'Invoice Number'])
        
        # Add table header
        sheet[f'A{start_row}'] = 'DETAILED GL UNMATCHED DATA'
        sheet[f'A{start_row}'].font = Font(bold=True, size=12, color='FF0000')
        start_row += 1
        
        # Add column headers
        sheet[f'A{start_row}'] = 'Invoice Number'
        sheet[f'B{start_row}'] = 'Entered Amount'
        sheet[f'C{start_row}'] = 'Status'
        
        # Style headers
        for cell in ['A', 'B', 'C']:
            sheet[f'{cell}{start_row}'].font = Font(bold=True)
            sheet[f'{cell}{start_row}'].alignment = Alignment(horizontal='center')
        
        start_row += 1
        
        # Fill data - show all individual line items
        data_start_row = start_row
        for idx, row in gl_detailed.iterrows():
            sheet[f'A{start_row}'] = str(row['Invoice Number'])[:80]  # Limit description length
            sheet[f'B{start_row}'] = row['Entered Amount']
            start_row += 1
        
        # Add borders
        add_detailed_table_borders(sheet, len(gl_detailed), data_start_row - 1, ['A', 'B', 'C'])
        
        # Compare and mark status for detailed GL table
        compare_detailed_gl_status(sheet, gl_detailed, data_start_row)
        
        return gl_detailed, data_start_row  # Return the data and data start row for cross-comparison
        
    except Exception as e:
        print(f"Error creating detailed GL table: {str(e)}")
        return pd.DataFrame(), start_row

def create_detailed_ibs_table(sheet, cost_extract_filtered, ibs_unmatched_dates, start_row):
    """
    Create detailed IBS table below main IBS table.
    Shows individual line items only for unmatched dates.
    """
    try:
        # Filter IBS data to show only line items for unmatched dates
        if ibs_unmatched_dates:
            # Convert dates to datetime for comparison
            cost_extract_copy = cost_extract_filtered.copy()
            cost_extract_copy['Invoice Date'] = pd.to_datetime(cost_extract_copy['Invoice Date'], errors='coerce')
            
            # Filter to show only line items for unmatched dates
            ibs_detailed = cost_extract_copy[cost_extract_copy['Invoice Date'].isin(ibs_unmatched_dates)]
        else:
            ibs_detailed = pd.DataFrame()  # Empty dataframe if no unmatched dates
        
        if len(ibs_detailed) == 0:
            # Still create the table header even if no data
            sheet[f'D{start_row}'] = 'DETAILED IBS UNMATCHED DATA'
            sheet[f'D{start_row}'].font = Font(bold=True, size=12, color='FF0000')
            start_row += 1
            sheet[f'D{start_row}'] = 'Invoice Number'
            sheet[f'E{start_row}'] = 'Invoice Amount'
            start_row += 1
            sheet[f'D{start_row}'] = 'No unmatched data found'
            return
        
        # Sort by Invoice Date first, then by Invoice Number
        ibs_detailed = ibs_detailed.copy()
        ibs_detailed['Invoice Date'] = pd.to_datetime(ibs_detailed['Invoice Date'], errors='coerce')
        ibs_detailed = ibs_detailed.sort_values(['Invoice Date', 'Invoice Number'])
        
        # Add table header
        sheet[f'D{start_row}'] = 'DETAILED IBS UNMATCHED DATA'
        sheet[f'D{start_row}'].font = Font(bold=True, size=12, color='FF0000')
        start_row += 1
        
        # Add column headers
        sheet[f'D{start_row}'] = 'Invoice Number'
        sheet[f'E{start_row}'] = 'Invoice Amount'
        sheet[f'F{start_row}'] = 'Status'
        
        # Style headers
        for cell in ['D', 'E', 'F']:
            sheet[f'{cell}{start_row}'].font = Font(bold=True)
            sheet[f'{cell}{start_row}'].alignment = Alignment(horizontal='center')
        
        start_row += 1
        
        # Fill data - show all individual line items
        data_start_row = start_row
        for idx, row in ibs_detailed.iterrows():
            sheet[f'D{start_row}'] = str(row['Invoice Number'])[:20]  # Limit invoice number length
            sheet[f'E{start_row}'] = row['Invoice Amount']
            start_row += 1
        
        # Add borders
        add_detailed_table_borders(sheet, len(ibs_detailed), data_start_row - 1, ['D', 'E', 'F'])
        
        # Compare and mark status for detailed IBS table
        compare_detailed_ibs_status(sheet, ibs_detailed, data_start_row)
        
        return ibs_detailed, data_start_row  # Return the data and data start row for cross-comparison
        
    except Exception as e:
        print(f"Error creating detailed IBS table: {str(e)}")
        return pd.DataFrame(), start_row

def cross_compare_detailed_tables(sheet, gl_detailed_data, ibs_detailed_data, gl_data_start_row, ibs_data_start_row):
    """
    Cross-compare detailed GL and IBS tables and update status columns.
    """
    try:
        if len(gl_detailed_data) == 0 and len(ibs_detailed_data) == 0:
            return
        
        # Get all amounts from both detailed tables
        gl_amounts = set()
        ibs_amounts = set()
        
        if len(gl_detailed_data) > 0:
            gl_amounts = set(abs(float(row['Entered Amount'])) for _, row in gl_detailed_data.iterrows())
        
        if len(ibs_detailed_data) > 0:
            ibs_amounts = set(abs(float(row['Invoice Amount'])) for _, row in ibs_detailed_data.iterrows())
        
        # Find matched and unmatched amounts
        matched_amounts = gl_amounts.intersection(ibs_amounts)
        gl_unmatched = gl_amounts - ibs_amounts
        ibs_unmatched = ibs_amounts - gl_amounts
        
        # Update GL detailed table status
        if len(gl_detailed_data) > 0:
            for i, (idx, row) in enumerate(gl_detailed_data.iterrows()):
                amount = abs(float(row['Entered Amount']))
                row_num = gl_data_start_row + i  # Use GL data start row
                if amount in matched_amounts:
                    sheet[f'C{row_num}'] = 'MATCHED'
                    sheet[f'C{row_num}'].font = Font(color='008000', bold=True)  # Green color
                else:
                    sheet[f'C{row_num}'] = 'UNMATCHED'
                    sheet[f'C{row_num}'].font = Font(color='FF0000', bold=True)  # Red color
        
        # Update IBS detailed table status
        if len(ibs_detailed_data) > 0:
            for i, (idx, row) in enumerate(ibs_detailed_data.iterrows()):
                amount = abs(float(row['Invoice Amount']))
                row_num = ibs_data_start_row + i  # Use IBS data start row
                if amount in matched_amounts:
                    sheet[f'F{row_num}'] = 'MATCHED'
                    sheet[f'F{row_num}'].font = Font(color='008000', bold=True)  # Green color
                else:
                    sheet[f'F{row_num}'] = 'UNMATCHED'
                    sheet[f'F{row_num}'].font = Font(color='FF0000', bold=True)  # Red color
        
    except Exception as e:
        print(f"Error in cross-comparison: {str(e)}")

def compare_detailed_gl_status(sheet, gl_detailed, start_row):
    """
    Initial status marking for GL detailed table (will be updated by cross-comparison).
    """
    try:
        # Initially mark all as UNMATCHED (will be updated by cross-comparison)
        for i, (idx, row) in enumerate(gl_detailed.iterrows()):
            row_num = start_row + i  # Data rows start at start_row
            sheet[f'C{row_num}'] = 'UNMATCHED'
            sheet[f'C{row_num}'].font = Font(color='FF0000', bold=True)  # Red color
        
    except Exception as e:
        print(f"Error comparing detailed GL status: {str(e)}")

def compare_detailed_ibs_status(sheet, ibs_detailed, start_row):
    """
    Initial status marking for IBS detailed table (will be updated by cross-comparison).
    """
    try:
        # Initially mark all as UNMATCHED (will be updated by cross-comparison)
        for i, (idx, row) in enumerate(ibs_detailed.iterrows()):
            row_num = start_row + i  # Data rows start at start_row
            sheet[f'F{row_num}'] = 'UNMATCHED'
            sheet[f'F{row_num}'].font = Font(color='FF0000', bold=True)  # Red color
        
    except Exception as e:
        print(f"Error comparing detailed IBS status: {str(e)}")

def add_detailed_table_borders(sheet, rows, start_row, columns):
    """
    Add borders to the detailed unmatched tables.
    """
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add borders to detailed table
    for row in range(start_row, start_row + rows + 1):  # +1 for header row
        for col in columns:
            sheet[f'{col}{row}'].border = thin_border

def add_table_borders(sheet, gl_rows, ibs_rows):
    """
    Add borders to the GL and IBS tables.
    """
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # GL Table borders (including Status column C)
    for row in range(2, gl_rows + 3):  # Start from header row
        for col in ['A', 'B', 'C']:
            sheet[f'{col}{row}'].border = thin_border
    
    # IBS Table borders (including Status column F)
    for row in range(2, ibs_rows + 3):  # Start from header row
        for col in ['D', 'E', 'F']:
            sheet[f'{col}{row}'].border = thin_border

def validate_unmatched_data(sheet, gl_detailed_data, ibs_detailed_data, cost_extract_data, row_label):
    """
    Validate unmatched data by:
    1. Sum all invoice amounts for invoices with UNMATCHED status in both detailed tables (Variable A)
    2. Filter Cost extract report by sheet name and unmatched invoice amounts
    3. Sum Total Of Invoice Indirect Tax Charges from filtered data (Variable B)
    
    Returns tuple (variable_A, variable_B, validation_summary)
    """
    try:
        variable_A = 0
        variable_B = 0
        validation_summary = {}
        
        # Step 1: Get unmatched invoices from both detailed tables by checking status column
        gl_unmatched_invoices = set()
        ibs_unmatched_invoices = set()
        
        # Get unmatched invoices from GL detailed table
        if len(gl_detailed_data) > 0:
            # Find the starting row of GL detailed data in the sheet
            gl_start_row = None
            for row_num in range(1, 100):  # Search for GL detailed table header
                if sheet[f'A{row_num}'].value == 'DETAILED GL UNMATCHED DATA':
                    gl_start_row = row_num + 2  # Data starts 2 rows below header
                    break
            
            if gl_start_row:
                for i, (idx, row) in enumerate(gl_detailed_data.iterrows()):
                    # Check status in column C
                    status_cell = sheet[f'C{gl_start_row + i}']
                    if status_cell.value == 'UNMATCHED':
                        invoice_number = str(row['Invoice Number'])
                        entered_amount = float(row['Entered Amount'])
                        gl_unmatched_invoices.add((invoice_number, entered_amount))
                        variable_A += entered_amount
        
        # Get unmatched invoices from IBS detailed table
        if len(ibs_detailed_data) > 0:
            # Find the starting row of IBS detailed data in the sheet
            ibs_start_row = None
            for row_num in range(1, 100):  # Search for IBS detailed table header
                if sheet[f'D{row_num}'].value == 'DETAILED IBS UNMATCHED DATA':
                    ibs_start_row = row_num + 2  # Data starts 2 rows below header
                    break
            
            if ibs_start_row:
                for i, (idx, row) in enumerate(ibs_detailed_data.iterrows()):
                    # Check status in column F
                    status_cell = sheet[f'F{ibs_start_row + i}']
                    if status_cell.value == 'UNMATCHED':
                        invoice_number = str(row['Invoice Number'])
                        invoice_amount = float(row['Invoice Amount'])
                        ibs_unmatched_invoices.add((invoice_number, invoice_amount))
                        variable_A += invoice_amount
        
        # Step 2: Filter Cost extract report by sheet name (row_label) and unmatched invoice amounts
        cost_extract_filtered = cost_extract_data[cost_extract_data['Concatenate'] == row_label]
        
        # Get all unmatched invoice numbers for filtering
        all_unmatched_invoices = set()
        for invoice_num, _ in gl_unmatched_invoices:
            all_unmatched_invoices.add(invoice_num)
        for invoice_num, _ in ibs_unmatched_invoices:
            all_unmatched_invoices.add(invoice_num)
        
        # Filter by unmatched invoice numbers
        cost_extract_unmatched = pd.DataFrame()
        if len(all_unmatched_invoices) > 0:
            # Convert invoice numbers to string for comparison
            cost_extract_filtered_copy = cost_extract_filtered.copy()
            cost_extract_filtered_copy['Invoice Number'] = cost_extract_filtered_copy['Invoice Number'].astype(str)
            unmatched_invoice_filter = cost_extract_filtered_copy['Invoice Number'].isin(all_unmatched_invoices)
            cost_extract_unmatched = cost_extract_filtered_copy[unmatched_invoice_filter]
            
            # Step 3: Sum Total Of Invoice Indirect Tax Charges
            if 'Total Of Invoice Indirect Tax Charges' in cost_extract_unmatched.columns:
                variable_B = cost_extract_unmatched['Total Of Invoice Indirect Tax Charges'].sum()
            else:
                print(f"Warning: 'Total Of Invoice Indirect Tax Charges' column not found in Cost extract report")
                variable_B = 0
        
        # Create validation summary
        validation_summary = {
            'gl_unmatched_count': len(gl_unmatched_invoices),
            'ibs_unmatched_count': len(ibs_unmatched_invoices),
            'total_unmatched_invoices': len(all_unmatched_invoices),
            'variable_A': variable_A,
            'variable_B': variable_B,
            'cost_extract_filtered_count': len(cost_extract_filtered),
            'cost_extract_unmatched_count': len(cost_extract_unmatched)
        }
        
        print(f"Validation for {row_label}:")
        print(f"  Variable A (Sum of unmatched amounts): {variable_A}")
        print(f"  Variable B (Sum of Indirect Tax Charges): {variable_B}")
        print(f"  GL Unmatched invoices: {len(gl_unmatched_invoices)}")
        print(f"  IBS Unmatched invoices: {len(ibs_unmatched_invoices)}")
        
        return variable_A, variable_B, validation_summary, cost_extract_unmatched
        
    except Exception as e:
        print(f"Error in validation: {str(e)}")
        return 0, 0, {}, pd.DataFrame()

def create_table1(sheet, cost_extract_unmatched_data, start_row=1):
    """
    Create Table 1 in column H (8th column) starting from row 1.
    Table contains Invoice Number and VAT/TAX (Total Of Invoice Indirect Tax Charges).
    """
    try:
        # Table 1 starts at column H (8th column), row 1
        table_start_col = 'H'
        table_start_row = start_row
        
        # Add table header
        sheet[f'{table_start_col}{table_start_row}'] = 'Table 1'
        sheet[f'{table_start_col}{table_start_row}'].font = Font(bold=True, size=12, color='0000FF')
        table_start_row += 1
        
        # Add column headers
        sheet[f'{table_start_col}{table_start_row}'] = 'Invoice Number'
        sheet[f'I{table_start_row}'] = 'VAT/TAX'
        
        # Style headers
        for col in [table_start_col, 'I']:
            sheet[f'{col}{table_start_row}'].font = Font(bold=True)
            sheet[f'{col}{table_start_row}'].alignment = Alignment(horizontal='center')
        
        table_start_row += 1
        
        # Fill data
        if len(cost_extract_unmatched_data) > 0:
            for idx, row in cost_extract_unmatched_data.iterrows():
                sheet[f'{table_start_col}{table_start_row}'] = str(row['Invoice Number'])
                sheet[f'I{table_start_row}'] = row['Total Of Invoice Indirect Tax Charges']
                table_start_row += 1
        else:
            sheet[f'{table_start_col}{table_start_row}'] = 'No data available'
        
        # Add borders to Table 1
        add_table1_borders(sheet, len(cost_extract_unmatched_data), start_row, [table_start_col, 'I'])
        
        # Auto-adjust column widths for Table 1
        sheet.column_dimensions[table_start_col].width = 20
        sheet.column_dimensions['I'].width = 15
        
        print(f"Created Table 1 with {len(cost_extract_unmatched_data)} records")
        
    except Exception as e:
        print(f"Error creating Table 1: {str(e)}")

def add_table1_borders(sheet, rows, start_row, columns):
    """
    Add borders to Table 1.
    """
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add borders to Table 1 (including header row)
    for row in range(start_row, start_row + rows + 2):  # +2 for header and title rows
        for col in columns:
            sheet[f'{col}{row}'].border = thin_border

def find_unmatched_invoices_not_in_both_tables(gl_detailed_data, ibs_detailed_data, sheet):
    """
    Find invoices with unmatched status that are not present in both detailed tables.
    Returns list of unmatched invoice numbers.
    """
    try:
        unmatched_invoices = set()
        
        # Get all invoice numbers from both detailed tables
        gl_invoices = set()
        ibs_invoices = set()
        
        if len(gl_detailed_data) > 0:
            gl_invoices = set(str(row['Invoice Number']) for _, row in gl_detailed_data.iterrows())
        
        if len(ibs_detailed_data) > 0:
            ibs_invoices = set(str(row['Invoice Number']) for _, row in ibs_detailed_data.iterrows())
        
        # Find invoices that are in one table but not in both
        gl_only = gl_invoices - ibs_invoices
        ibs_only = ibs_invoices - gl_invoices
        
        # Check status for GL-only invoices
        if len(gl_detailed_data) > 0:
            gl_start_row = None
            for row_num in range(1, 100):
                if sheet[f'A{row_num}'].value == 'DETAILED GL UNMATCHED DATA':
                    gl_start_row = row_num + 2
                    break
            
            if gl_start_row:
                for i, (idx, row) in enumerate(gl_detailed_data.iterrows()):
                    invoice_number = str(row['Invoice Number'])
                    if invoice_number in gl_only:
                        status_cell = sheet[f'C{gl_start_row + i}']
                        if status_cell.value == 'UNMATCHED':
                            unmatched_invoices.add(invoice_number)
        
        # Check status for IBS-only invoices
        if len(ibs_detailed_data) > 0:
            ibs_start_row = None
            for row_num in range(1, 100):
                if sheet[f'D{row_num}'].value == 'DETAILED IBS UNMATCHED DATA':
                    ibs_start_row = row_num + 2
                    break
            
            if ibs_start_row:
                for i, (idx, row) in enumerate(ibs_detailed_data.iterrows()):
                    invoice_number = str(row['Invoice Number'])
                    if invoice_number in ibs_only:
                        status_cell = sheet[f'F{ibs_start_row + i}']
                        if status_cell.value == 'UNMATCHED':
                            unmatched_invoices.add(invoice_number)
        
        print(f"Found {len(unmatched_invoices)} invoices with unmatched status not present in both tables")
        return list(unmatched_invoices)
        
    except Exception as e:
        print(f"Error finding unmatched invoices not in both tables: {str(e)}")
        return []

def get_entered_net_from_gl(gl_data, unmatched_invoices, row_label):
    """
    Get Entered Net amounts from GL sheet for the unmatched invoices.
    Returns list of tuples (invoice_number, entered_net_amount).
    """
    try:
        rts_data = []
        
        # Filter GL data by row_label
        gl_filtered = gl_data[gl_data['Concatenate'] == row_label]
        
        # Convert invoice numbers to string for comparison
        gl_filtered_copy = gl_filtered.copy()
        gl_filtered_copy['Invoice Number'] = gl_filtered_copy['Invoice Number'].astype(str)
        
        # Filter by unmatched invoice numbers
        if len(unmatched_invoices) > 0:
            unmatched_invoice_filter = gl_filtered_copy['Invoice Number'].isin(unmatched_invoices)
            gl_unmatched = gl_filtered_copy[unmatched_invoice_filter]
            
            # Get Entered Net amounts
            if 'Entered Net' in gl_unmatched.columns:
                for idx, row in gl_unmatched.iterrows():
                    invoice_number = str(row['Invoice Number'])
                    entered_net = row['Entered Net']
                    rts_data.append((invoice_number, entered_net))
            else:
                print(f"Warning: 'Entered Net' column not found in GL data")
        
        print(f"Found {len(rts_data)} RTS records from GL data")
        return rts_data
        
    except Exception as e:
        print(f"Error getting Entered Net from GL: {str(e)}")
        return []

def create_rts_table(sheet, rts_data, start_row=1):
    """
    Create RTS table in column M (13th column) starting from row 1.
    Table contains Invoice Number and RTS (Entered Net amounts).
    """
    try:
        # RTS table starts at column M (13th column), row 1
        table_start_col = 'M'
        table_start_row = start_row
        
        # Add table header
        sheet[f'{table_start_col}{table_start_row}'] = 'RTS Table'
        sheet[f'{table_start_col}{table_start_row}'].font = Font(bold=True, size=12, color='0000FF')
        table_start_row += 1
        
        # Add column headers
        sheet[f'{table_start_col}{table_start_row}'] = 'Invoice Number'
        sheet[f'N{table_start_row}'] = 'RTS'
        
        # Style headers
        for col in [table_start_col, 'N']:
            sheet[f'{col}{table_start_row}'].font = Font(bold=True)
            sheet[f'{col}{table_start_row}'].alignment = Alignment(horizontal='center')
        
        table_start_row += 1
        
        # Fill data
        if len(rts_data) > 0:
            for invoice_number, entered_net in rts_data:
                sheet[f'{table_start_col}{table_start_row}'] = str(invoice_number)
                sheet[f'N{table_start_row}'] = entered_net
                table_start_row += 1
        else:
            sheet[f'{table_start_col}{table_start_row}'] = 'No data available'
        
        # Add borders to RTS table
        add_rts_table_borders(sheet, len(rts_data), start_row, [table_start_col, 'N'])
        
        # Auto-adjust column widths for RTS table
        sheet.column_dimensions[table_start_col].width = 20
        sheet.column_dimensions['N'].width = 15
        
        print(f"Created RTS table with {len(rts_data)} records")
        
    except Exception as e:
        print(f"Error creating RTS table: {str(e)}")

def add_rts_table_borders(sheet, rows, start_row, columns):
    """
    Add borders to RTS table.
    """
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add borders to RTS table (including header row)
    for row in range(start_row, start_row + rows + 2):  # +2 for header and title rows
        for col in columns:
            sheet[f'{col}{row}'].border = thin_border

def compare_variables_and_create_table1(sheet, variable_A, variable_B, cost_extract_unmatched_data):
    """
    Compare variables A and B, and create Table 1 if they are equal.
    """
    try:
        # Compare A and B (with small tolerance for floating point comparison)
        tolerance = 0.01
        if abs(variable_A - variable_B) <= tolerance:
            print(f"Variables A and B are equal (A={variable_A}, B={variable_B}). Creating Table 1...")
            create_table1(sheet, cost_extract_unmatched_data, start_row=1)
            return True
        else:
            print(f"Variables A and B are not equal (A={variable_A}, B={variable_B}). Table 1 not created.")
            return False
    except Exception as e:
        print(f"Error comparing variables: {str(e)}")
        return False

def add_validation_summary_to_sheet(sheet, variable_A, variable_B, validation_summary, start_row):
    """
    Add validation summary to the sheet below the detailed tables.
    """
    try:
        # Add validation summary section
        sheet[f'A{start_row}'] = 'VALIDATION SUMMARY'
        sheet[f'A{start_row}'].font = Font(bold=True, size=12, color='0000FF')
        start_row += 1
        
        sheet[f'A{start_row}'] = f'Variable A (Sum of Unmatched Amounts): {variable_A:,.2f}'
        sheet[f'A{start_row}'].font = Font(bold=True)
        start_row += 1
        
        sheet[f'A{start_row}'] = f'Variable B (Sum of Indirect Tax Charges): {variable_B:,.2f}'
        sheet[f'A{start_row}'].font = Font(bold=True)
        start_row += 1
        
        # Add comparison result
        tolerance = 0.01
        if abs(variable_A - variable_B) <= tolerance:
            sheet[f'A{start_row}'] = 'COMPARISON RESULT: A = B ✓ (Table 1 Created)'
            sheet[f'A{start_row}'].font = Font(bold=True, color='008000')
        else:
            sheet[f'A{start_row}'] = f'COMPARISON RESULT: A ≠ B (Difference: {abs(variable_A - variable_B):,.2f})'
            sheet[f'A{start_row}'].font = Font(bold=True, color='FF0000')
        start_row += 1
        
        start_row += 1  # Empty line
        
        # Add detailed breakdown
        sheet[f'A{start_row}'] = 'Detailed Breakdown:'
        sheet[f'A{start_row}'].font = Font(bold=True)
        start_row += 1
        
        sheet[f'A{start_row}'] = f'GL Unmatched Invoices: {validation_summary.get("gl_unmatched_count", 0)}'
        start_row += 1
        
        sheet[f'A{start_row}'] = f'IBS Unmatched Invoices: {validation_summary.get("ibs_unmatched_count", 0)}'
        start_row += 1
        
        sheet[f'A{start_row}'] = f'Total Unmatched Invoices: {validation_summary.get("total_unmatched_invoices", 0)}'
        start_row += 1
        
        sheet[f'A{start_row}'] = f'Cost Extract Filtered Records: {validation_summary.get("cost_extract_filtered_count", 0)}'
        start_row += 1
        
        sheet[f'A{start_row}'] = f'Cost Extract Unmatched Records: {validation_summary.get("cost_extract_unmatched_count", 0)}'
        start_row += 1
        
        # Add RTS table information if available
        if hasattr(sheet, '_validation_results') and 'rts_data' in sheet._validation_results:
            rts_count = len(sheet._validation_results['rts_data'])
            sheet[f'A{start_row}'] = f'RTS Table Records: {rts_count}'
            start_row += 1
        
    except Exception as e:
        print(f"Error adding validation summary to sheet: {str(e)}")

def clean_sheet_name(name):
    """
    Clean the row label to make it a valid Excel sheet name.
    Excel sheet names cannot contain: \ / ? * [ ]
    Also limit length to 31 characters (Excel limitation)
    """
    # Replace invalid characters with underscores
    invalid_chars = ['\\', '/', '?', '*', '[', ']', ':', ';']
    for char in invalid_chars:
        name = str(name).replace(char, '_')
    
    # Remove leading/trailing spaces and dots
    name = name.strip('. ')
    
    # Limit length to 31 characters
    if len(name) > 31:
        name = name[:31]
    
    # Ensure it's not empty
    if not name:
        name = "Sheet"
    
    return name

def run_complete_reconciliation():
    """
    Run the complete reconciliation process including variance analysis.
    This ensures everything is written to the same main file.
    """
    print("CP-ESCOA Cost Extract Reconciliation - Complete Process")
    print("=" * 60)
    
    # The main reconciliation process runs automatically when the script is imported/run
    # The main_path variable is set during the initial file processing
    
    # After the main reconciliation is complete, run variance analysis
    if 'main_path' in globals():
        print(f"\nMain reconciliation completed. Now running variance analysis...")
        print(f"Using main file: {main_path}")
        create_variance_sheets(main_path)
    else:
        print("Error: Main reconciliation process did not complete successfully.")
        return False
    
    print("\nComplete reconciliation process finished!")
    return True

if __name__ == "__main__":
    print("CP-ESCOA Cost Extract Reconciliation - Variance Analysis")
    print("=" * 60)
    
    # Check if main_path is available from the main reconciliation process
    if 'main_path' in globals():
        print(f"Using main file: {main_path}")
        create_variance_sheets(main_path)
    else:
        print("Main reconciliation not run yet. Running variance analysis with default files...")
        create_variance_sheets()
    
    print("\nProcess completed!")
